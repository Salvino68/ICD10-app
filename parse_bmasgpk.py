#!/usr/bin/env python3
"""
BMASGPK PDF Parser - Extrahiert ICD-10 Codes aus dem BMASGPK Leistungskatalog PDF
und konvertiert sie zu XLSX für die ICD-10 Generator App
"""

import re
import requests
import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import Font, PatternFill
from pathlib import Path
import sys

# URL des PDF vom Sozialministerium
PDF_URL = "https://www.sozialministerium.gv.at/dam/jcr:c4a10158-5a38-4dcb-8771-9134d7b6d2f6/LEISTUNGSKATALOG%20BMASGPK%202026.pdf"
OUTPUT_FILE = "DIAGLIST_2026.xlsx"

def download_pdf(url):
    """PDF vom Sozialministerium herunterladen"""
    print(f"📥 Lade PDF herunter: {url}")
    try:
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        return response.content
    except Exception as e:
        print(f"❌ Fehler beim Download: {e}")
        return None

def parse_pdf(pdf_data):
    """PDF parsen und Codes extrahieren"""
    print("📖 Parsing PDF...")
    codes = []
    
    try:
        pdf_document = fitz.open(stream=pdf_data, filetype="pdf")
        
        # Regex Patterns für verschiedene Code-Formate
        # ICD-10 Format: A12, B45, Z99, etc.
        icd_pattern = r'^([A-Z]\d{2}(?:\.\d{1,3})?)\s+(.+)$'
        
        current_chapter = ""
        
        for page_num in range(len(pdf_document)):
            page = pdf_document[page_num]
            text = page.get_text()
            lines = text.split('\n')
            
            for line in lines:
                line = line.strip()
                
                # Kapitel erkennen (zB "01 Nervensystem")
                if re.match(r'^\d{2}\s+', line):
                    current_chapter = line[:80]
                    continue
                
                # ICD-10 Codes erkennen
                match = re.match(icd_pattern, line)
                if match:
                    code = match.group(1).upper()
                    description = match.group(2).strip()
                    
                    # Filter: nur ICD-10 Codes (A-Z, gefolgt von Zahlen)
                    if re.match(r'^[A-Z]\d', code) and len(description) > 3:
                        codes.append({
                            'code': code,
                            'bez': description,
                            'kap': current_chapter
                        })
        
        pdf_document.close()
        
    except Exception as e:
        print(f"⚠️  PDF Parse Fehler: {e}")
        print("   Fallback: Verwende Text-Extraktion...")
        # Fallback: Text direkt extrahieren
        codes = parse_text_fallback(pdf_data)
    
    return codes

def parse_text_fallback(pdf_data):
    """Fallback: Direktes Text-Parsing wenn PDF-Parsing fehlschlägt"""
    codes = []
    try:
        pdf_document = fitz.open(stream=pdf_data, filetype="pdf")
        all_text = ""
        for page in pdf_document:
            all_text += page.get_text()
        pdf_document.close()
        
        lines = all_text.split('\n')
        current_chapter = ""
        
        for line in lines:
            line = line.strip()
            
            # Kapitel
            if re.match(r'^(\d{2,3})\s+', line) and len(line) > 5:
                current_chapter = line[:80]
            
            # Codes suchen
            match = re.match(r'^([A-Z]\d{2}(?:\.\d{1,3})?)\s+(.+)$', line)
            if match:
                code = match.group(1).upper()
                description = match.group(2).strip()
                if len(description) > 3:
                    codes.append({
                        'code': code,
                        'bez': description,
                        'kap': current_chapter
                    })
    except Exception as e:
        print(f"⚠️  Fallback auch fehlgeschlagen: {e}")
    
    return codes

def create_xlsx(codes, output_file):
    """Codes zu XLSX speichern"""
    print(f"📊 Erstelle XLSX mit {len(codes)} Codes...")
    
    # Duplikate entfernen
    seen = set()
    unique_codes = []
    for c in codes:
        key = (c['code'], c['bez'])
        if key not in seen:
            seen.add(key)
            unique_codes.append(c)
    
    # Excel Workbook erstellen
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ICD-10"
    
    # Header
    headers = ['Code', 'Bezeichnung', 'Kapitel']
    header_fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    
    # Daten
    for row_num, code_data in enumerate(unique_codes, 2):
        ws.cell(row=row_num, column=1).value = code_data['code']
        ws.cell(row=row_num, column=2).value = code_data['bez']
        ws.cell(row=row_num, column=3).value = code_data['kap']
    
    # Spaltenbreiten
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 35
    
    # Speichern
    wb.save(output_file)
    print(f"✅ Gespeichert: {output_file}")

def main():
    print("=" * 60)
    print("🏥 BMASGPK PDF Parser - ICD-10 2026")
    print("=" * 60)
    
    # PDF herunterladen
    pdf_data = download_pdf(PDF_URL)
    if not pdf_data:
        print("❌ Abbruch: PDF konnte nicht heruntergeladen werden")
        return False
    
    # PDF parsen
    codes = parse_pdf(pdf_data)
    
    if not codes or len(codes) < 100:
        print(f"⚠️  Nur {len(codes)} Codes gefunden - Manueller Import empfohlen")
    
    # XLSX erstellen
    create_xlsx(codes, OUTPUT_FILE)
    
    print("=" * 60)
    print(f"✨ Fertig! {len(codes)} ICD-10 Codes verarbeitet")
    print(f"📁 Datei: {OUTPUT_FILE}")
    print("=" * 60)
    
    return True

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
